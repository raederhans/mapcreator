"""Build the WGI real-source admin0 thematic layer from a local source cache."""
from __future__ import annotations

import csv
import hashlib
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_GENERATED_AT = "2026-06-22T00:00:00Z"

WGI_LAYER_ID = "political_wgi_state_capacity_v1"
WGI_SELECTED_YEAR = 2024
WGI_SOURCE_TITLE = "WGI 2025 Revision: Governance Estimates and Absolute Scores (1996-2024)"
WGI_SOURCE_VERSION = "7"
WGI_SOURCE_URL = "https://datacatalog.worldbank.org/search/dataset/0038026/worldwide-governance-indicators"
WGI_SOURCE_PACKAGE_URL = "https://datacatalogfiles.worldbank.org/ddh-published/0038026/DR0095947/wgidataset_with_sourcedata-2025.xlsx"
WGI_METADATA_UPDATED_AT = "2026-03-18"
WGI_PACKAGE_PUBLISHED_AT = "2026-03-11"
WGI_LICENSE = "Creative Commons Attribution 4.0"
WGI_CITATION = "World Bank Worldwide Governance Indicators Data Catalog entry."

WGI_SOURCE_CACHE_RELATIVE_PATH = (
    ".runtime/source-cache/thematic/wgi/"
    "WGI_2025_Revision_Governance_Estimates_and_Absolute_Scores_1996_2024.xlsx"
)
DEFAULT_WGI_SOURCE_CACHE_PATH = REPO_ROOT / WGI_SOURCE_CACHE_RELATIVE_PATH

WGI_MANIFEST_RELATIVE_PATH = "thematic_layers/political/wgi_state_capacity_v1/manifest.json"
WGI_METRICS_RELATIVE_PATH = "thematic_layers/political/wgi_state_capacity_v1/metrics.admin0.json"
WGI_AUDIT_RELATIVE_PATH = "thematic_layers/political/wgi_state_capacity_v1/build_audit.json"
WGI_RECIPE_RELATIVE_PATH = "thematic_layers/source_recipes/wgi_state_capacity_v1.manual.json"

WGI_GOVERNMENT_EFFECTIVENESS_METRIC_ID = "wgi_government_effectiveness_score_0_100"
WGI_RULE_OF_LAW_METRIC_ID = "wgi_rule_of_law_score_0_100"
WGI_COMPOSITE_METRIC_ID = "wgi_state_capacity_composite_0_100"
WGI_METRIC_IDS = (
    WGI_GOVERNMENT_EFFECTIVENESS_METRIC_ID,
    WGI_RULE_OF_LAW_METRIC_ID,
    WGI_COMPOSITE_METRIC_ID,
)

DIMENSION_SHEETS = {
    "ge": "government_effectiveness",
    "rl": "rule_of_law",
}
DIMENSION_TO_METRIC_ID = {
    "government_effectiveness": WGI_GOVERNMENT_EFFECTIVENESS_METRIC_ID,
    "rule_of_law": WGI_RULE_OF_LAW_METRIC_ID,
}
WGI_OFFICIAL_DIMENSIONS = {
    "government_effectiveness": "Government Effectiveness",
    "rule_of_law": "Rule of Law",
}

SOURCE_CODE_TO_ISO_A3 = {
    "ADO": "AND",
    "IMY": "IMN",
    "KSV": "XKX",
    "ROM": "ROU",
    "TMP": "TLS",
    "WBG": "PSE",
    "XKX": "XKX",
    "ZAR": "COD",
}

KNOWN_NON_ISO_ECONOMY_CODES = {
    "ANT",
    "CHI",
}

WGI_AGGREGATE_SOURCE_CODES = {
    "AFE",
    "AFW",
    "ARB",
    "CEB",
    "CSS",
    "EAP",
    "EAR",
    "EAS",
    "ECA",
    "ECS",
    "EMU",
    "EUU",
    "FCS",
    "HIC",
    "HPC",
    "IBD",
    "IBT",
    "IDA",
    "IDB",
    "IDX",
    "LAC",
    "LCN",
    "LDC",
    "LIC",
    "LMC",
    "LMY",
    "LTE",
    "MEA",
    "MIC",
    "MNA",
    "NAC",
    "OED",
    "OSS",
    "PRE",
    "PSS",
    "PST",
    "SAS",
    "SSA",
    "SSF",
    "SST",
    "TEA",
    "TEC",
    "TLA",
    "TMN",
    "TSA",
    "TSS",
    "UMC",
    "WLD",
}


@dataclass(frozen=True)
class WgiSourceSignature:
    path: Path
    repo_relative_path: str
    size_bytes: int
    sha256: str


@dataclass(frozen=True)
class WgiObservation:
    source_name: str
    source_code: str
    join_key: str
    dimension: str
    year: int
    score_0_100: float | None
    score_standard_error: float | None
    score_confidence_interval_lower_90: float | None
    score_confidence_interval_upper_90: float | None
    estimate: float | None
    estimate_standard_error: float | None
    estimate_confidence_interval_lower_90: float | None
    estimate_confidence_interval_upper_90: float | None
    number_of_sources: int | None
    source_row_ref: str


@dataclass(frozen=True)
class DroppedSourceRow:
    source_name: str
    source_code: str
    dimension: str
    year: int | None
    reason: str
    source_row_ref: str


def data_url(relative_path: str) -> str:
    return f"data/{relative_path}"


def source_signature(path: Path) -> WgiSourceSignature:
    resolved = path.resolve()
    if not resolved.is_file():
        raise FileNotFoundError(f"WGI source cache is missing: {resolved}")
    digest = hashlib.sha256()
    with resolved.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    try:
        repo_relative_path = resolved.relative_to(REPO_ROOT).as_posix()
    except ValueError:
        repo_relative_path = str(resolved)
    return WgiSourceSignature(
        path=resolved,
        repo_relative_path=repo_relative_path,
        size_bytes=resolved.stat().st_size,
        sha256=digest.hexdigest(),
    )


def _normalize_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def _select_field(row: dict[str, Any], aliases: tuple[str, ...]) -> Any:
    normalized = {_normalize_header(key): value for key, value in row.items()}
    for alias in aliases:
        value = normalized.get(_normalize_header(alias))
        if value is not None:
            return value
    return None


def _normalize_code(value: object) -> str:
    return str(value or "").strip().upper()


def _normalize_name(value: object) -> str:
    return str(value or "").strip()


def _parse_year(value: object) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _parse_number(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        parsed = float(value)
        return round(parsed, 6) if math.isfinite(parsed) else None
    text = str(value).strip()
    if not text or text in {"..", "NA", "N/A", "nan"}:
        return None
    try:
        parsed = float(text)
    except ValueError:
        return None
    return round(parsed, 6) if math.isfinite(parsed) else None


def _parse_score(value: object) -> float | None:
    score = _parse_number(value)
    return score if score is not None and 0 <= score <= 100 else None


def _parse_float(value: object) -> float | None:
    return _parse_number(value)


def _parse_count(value: object) -> int | None:
    parsed = _parse_number(value)
    if parsed is None or parsed < 0 or not float(parsed).is_integer():
        return None
    return int(parsed)


def _is_truthy(value: object) -> bool:
    text = str(value or "").strip().lower()
    return text in {"1", "true", "yes", "y", "aggregate"}


def source_code_to_join_key(source_code: str) -> str | None:
    normalized = _normalize_code(source_code)
    if normalized in WGI_AGGREGATE_SOURCE_CODES or normalized in KNOWN_NON_ISO_ECONOMY_CODES:
        return None
    mapped = SOURCE_CODE_TO_ISO_A3.get(normalized)
    if mapped:
        return mapped
    if re.fullmatch(r"[A-Z]{3}", normalized):
        return normalized
    return None


def _row_to_observation(
    row: dict[str, Any],
    *,
    source_label: str,
    row_number: int,
    dimension: str,
    selected_year: int,
) -> tuple[WgiObservation | None, DroppedSourceRow | None]:
    source_name = _normalize_name(_select_field(row, ("Economy (name)", "economy_name", "name")))
    source_code = _normalize_code(_select_field(row, ("Economy (code)", "economy_code", "code")))
    year = _parse_year(_select_field(row, ("Year", "year")))
    score = _parse_score(_select_field(row, ("Governance score (0-100)", "score_0_100", "score")))
    score_standard_error = _parse_float(_select_field(row, ("Standard error (gov. score)", "score_standard_error")))
    score_confidence_interval_lower_90 = _parse_score(
        _select_field(row, ("Lower threshold (90% conf. int. score)", "score_confidence_interval_lower_90"))
    )
    score_confidence_interval_upper_90 = _parse_score(
        _select_field(row, ("Upper threshold (90% conf. int. score)", "score_confidence_interval_upper_90"))
    )
    estimate = _parse_float(
        _select_field(row, ("Governance estimate (approx. -2.5 to +2.5)", "governance_estimate"))
    )
    estimate_standard_error = _parse_float(_select_field(row, ("Standard error (estimate)", "estimate_standard_error")))
    estimate_confidence_interval_lower_90 = _parse_float(
        _select_field(row, ("Lower threshold (90% conf. int. estimate)", "estimate_confidence_interval_lower_90"))
    )
    estimate_confidence_interval_upper_90 = _parse_float(
        _select_field(row, ("Upper threshold (90% conf. int. estimate)", "estimate_confidence_interval_upper_90"))
    )
    number_of_sources = _parse_count(_select_field(row, ("Number of sources", "number_of_sources")))
    source_row_ref = f"{source_label}:row:{row_number}"

    if year != selected_year:
        return None, None
    if not source_code:
        return None, DroppedSourceRow(source_name, source_code, dimension, year, "missing_source_code", source_row_ref)
    if source_code in WGI_AGGREGATE_SOURCE_CODES or _is_truthy(row.get("is_aggregate")):
        return None, DroppedSourceRow(source_name, source_code, dimension, year, "aggregate_row", source_row_ref)
    join_key = source_code_to_join_key(source_code)
    if not join_key:
        return None, DroppedSourceRow(source_name, source_code, dimension, year, "unmatched_join_key", source_row_ref)
    return (
        WgiObservation(
            source_name=source_name,
            source_code=source_code,
            join_key=join_key,
            dimension=dimension,
            year=selected_year,
            score_0_100=score,
            score_standard_error=score_standard_error,
            score_confidence_interval_lower_90=score_confidence_interval_lower_90,
            score_confidence_interval_upper_90=score_confidence_interval_upper_90,
            estimate=estimate,
            estimate_standard_error=estimate_standard_error,
            estimate_confidence_interval_lower_90=estimate_confidence_interval_lower_90,
            estimate_confidence_interval_upper_90=estimate_confidence_interval_upper_90,
            number_of_sources=number_of_sources,
            source_row_ref=source_row_ref,
        ),
        None,
    )


def _read_csv_rows(path: Path) -> Iterable[tuple[str, dict[str, Any], int]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row_number, row in enumerate(reader, start=2):
            dimension_value = _normalize_header(_select_field(row, ("Governance dimension", "dimension")))
            dimension = ""
            if "government effectiveness" in dimension_value or dimension_value == "ge":
                dimension = "government_effectiveness"
            elif "rule of law" in dimension_value or dimension_value == "rl":
                dimension = "rule_of_law"
            yield dimension, row, row_number


def _read_xlsx_rows(path: Path) -> Iterable[tuple[str, dict[str, Any], int]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("openpyxl is required to ingest WGI .xlsx source caches") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name, dimension in DIMENSION_SHEETS.items():
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"WGI source cache is missing required sheet: {sheet_name}")
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(values_only=True)
            try:
                header = [str(value or "").strip() for value in next(rows)]
            except StopIteration:
                continue
            for row_number, values in enumerate(rows, start=2):
                row = {header[index]: values[index] if index < len(values) else None for index in range(len(header))}
                yield dimension, row, row_number
    finally:
        workbook.close()


def load_wgi_observations(
    source_path: Path,
    *,
    selected_year: int = WGI_SELECTED_YEAR,
) -> tuple[list[WgiObservation], list[DroppedSourceRow], list[DroppedSourceRow]]:
    suffix = source_path.suffix.lower()
    if suffix == ".csv":
        source_rows = _read_csv_rows(source_path)
    elif suffix == ".xlsx":
        source_rows = _read_xlsx_rows(source_path)
    else:
        raise ValueError(f"Unsupported WGI source cache format: {source_path.suffix}")

    observations: list[WgiObservation] = []
    aggregate_rows: list[DroppedSourceRow] = []
    unmatched_rows: list[DroppedSourceRow] = []
    seen: set[tuple[str, str]] = set()
    duplicates: list[str] = []

    for dimension, row, row_number in source_rows:
        if dimension not in DIMENSION_TO_METRIC_ID:
            continue
        observation, dropped = _row_to_observation(
            row,
            source_label=source_path.name,
            row_number=row_number,
            dimension=dimension,
            selected_year=selected_year,
        )
        if dropped:
            if dropped.reason == "aggregate_row":
                aggregate_rows.append(dropped)
            else:
                unmatched_rows.append(dropped)
            continue
        if observation is None:
            continue
        key = (observation.dimension, observation.join_key)
        if key in seen:
            duplicates.append(f"{observation.dimension}:{observation.join_key}:{observation.source_row_ref}")
            continue
        seen.add(key)
        observations.append(observation)

    if duplicates:
        raise ValueError("Duplicate WGI rows for the same dimension and join key: " + ", ".join(duplicates))
    return observations, aggregate_rows, unmatched_rows


def _missing_metric_payload(*, year: int, source_status: str, notes: str) -> dict[str, Any]:
    return {
        "raw_value": None,
        "normalized_value": None,
        "year": year,
        "unit": "score_0_100",
        "source_status": source_status,
        "notes": notes,
    }


def _uncertainty_payload(observation: WgiObservation) -> dict[str, Any]:
    return {
        "number_of_sources": observation.number_of_sources,
        "score_standard_error": observation.score_standard_error,
        "score_confidence_interval_90": {
            "lower": observation.score_confidence_interval_lower_90,
            "upper": observation.score_confidence_interval_upper_90,
        },
        "estimate": observation.estimate,
        "estimate_standard_error": observation.estimate_standard_error,
        "estimate_confidence_interval_90": {
            "lower": observation.estimate_confidence_interval_lower_90,
            "upper": observation.estimate_confidence_interval_upper_90,
        },
    }


def _observed_metric_payload(observation: WgiObservation) -> dict[str, Any]:
    if observation.score_0_100 is None:
        return _missing_metric_payload(
            year=observation.year,
            source_status="source_gap",
            notes=f"WGI {observation.dimension} row exists but score is empty.",
        ) | {
            "uncertainty": _uncertainty_payload(observation),
            "source_country_code": observation.source_code,
            "source_row_ref": observation.source_row_ref,
        }
    return {
        "raw_value": observation.score_0_100,
        "normalized_value": observation.score_0_100,
        "year": observation.year,
        "unit": "score_0_100",
        "source_status": "observed",
        "notes": f"WGI {observation.dimension} score for {WGI_SELECTED_YEAR}.",
        "uncertainty": _uncertainty_payload(observation),
        "source_country_code": observation.source_code,
        "source_row_ref": observation.source_row_ref,
    }


def _composite_metric_payload(
    ge_observation: WgiObservation | None,
    rl_observation: WgiObservation | None,
    *,
    selected_year: int,
) -> dict[str, Any]:
    ge_score_observed = ge_observation is not None and ge_observation.score_0_100 is not None
    rl_score_observed = rl_observation is not None and rl_observation.score_0_100 is not None
    if ge_score_observed and rl_score_observed and ge_observation and rl_observation:
        composite = round((ge_observation.score_0_100 + rl_observation.score_0_100) / 2.0, 6)
        return {
            "raw_value": composite,
            "normalized_value": composite,
            "year": selected_year,
            "unit": "score_0_100",
            "source_status": "observed",
            "notes": "Project-defined proxy: mean of WGI government effectiveness and rule of law scores.",
            "uncertainty": {
                "method": "not_computed",
                "reason": "Composite uncertainty is not inferred from the two source dimensions.",
            },
            "source_row_refs": {
                "government_effectiveness": ge_observation.source_row_ref,
                "rule_of_law": rl_observation.source_row_ref,
            },
        }
    source_status = "partial_source_gap" if ge_score_observed or rl_score_observed else "source_gap"
    return _missing_metric_payload(
        year=selected_year,
        source_status=source_status,
        notes="Composite requires both WGI government effectiveness and rule of law scores.",
    )


def build_admin_metrics_payload(
    observations: list[WgiObservation],
    *,
    selected_year: int = WGI_SELECTED_YEAR,
) -> dict[str, Any]:
    by_join_key: dict[str, dict[str, WgiObservation]] = {}
    names: dict[str, str] = {}
    source_codes: dict[str, set[str]] = {}
    source_row_refs: dict[str, dict[str, str]] = {}
    for observation in observations:
        by_join_key.setdefault(observation.join_key, {})[observation.dimension] = observation
        names.setdefault(observation.join_key, observation.source_name)
        source_codes.setdefault(observation.join_key, set()).add(observation.source_code)
        source_row_refs.setdefault(observation.join_key, {})[observation.dimension] = observation.source_row_ref

    features: list[dict[str, Any]] = []
    for join_key in sorted(by_join_key):
        ge_observation = by_join_key[join_key].get("government_effectiveness")
        rl_observation = by_join_key[join_key].get("rule_of_law")
        values = {
            WGI_GOVERNMENT_EFFECTIVENESS_METRIC_ID: (
                _observed_metric_payload(ge_observation)
                if ge_observation
                else _missing_metric_payload(
                    year=selected_year,
                    source_status="source_gap",
                    notes="WGI government effectiveness source row is missing.",
                )
            ),
            WGI_RULE_OF_LAW_METRIC_ID: (
                _observed_metric_payload(rl_observation)
                if rl_observation
                else _missing_metric_payload(
                    year=selected_year,
                    source_status="source_gap",
                    notes="WGI rule of law source row is missing.",
                )
            ),
            WGI_COMPOSITE_METRIC_ID: _composite_metric_payload(
                ge_observation,
                rl_observation,
                selected_year=selected_year,
            ),
        }
        missing_count = sum(1 for value in values.values() if value["raw_value"] is None)
        coverage_status = "complete"
        if missing_count == len(values):
            coverage_status = "missing"
        elif missing_count:
            coverage_status = "partial"
        features.append(
            {
                "join_key": join_key,
                "name": names.get(join_key, join_key),
                "coverage_status": coverage_status,
                "source_country_codes": sorted(source_codes.get(join_key, [])),
                "source_row_refs": source_row_refs.get(join_key, {}),
                "values": values,
            }
        )

    return {
        "schema_version": 1,
        "layer_id": WGI_LAYER_ID,
        "geography_level": "admin0",
        "join_key_type": "iso_a3",
        "metric_ids": list(WGI_METRIC_IDS),
        "features": features,
        "notes": [
            "Built from a local World Bank WGI source cache; the builder does not download by default.",
            "Join keys use explicit WGI economy code rules; unmatched rows stay in build audit.",
            "WGI standard errors and 90% confidence intervals are preserved per source metric.",
        ],
    }


def coverage_counts_for_admin(payload: dict[str, Any]) -> dict[str, int]:
    counts = {"features": 0, "complete": 0, "partial": 0, "missing": 0}
    for feature in payload["features"]:
        counts["features"] += 1
        status = str(feature.get("coverage_status") or "")
        if status in counts:
            counts[status] += 1
    return counts


def _source_row_payload(row: DroppedSourceRow) -> dict[str, Any]:
    return {
        "source_name": row.source_name,
        "source_code": row.source_code,
        "dimension": row.dimension,
        "year": row.year,
        "reason": row.reason,
        "source_row_ref": row.source_row_ref,
    }


def _metric_values(metrics_payload: dict[str, Any], metric_id: str) -> list[tuple[str, str, float]]:
    values: list[tuple[str, str, float]] = []
    for feature in metrics_payload["features"]:
        metric = feature["values"].get(metric_id, {})
        normalized = metric.get("normalized_value")
        if isinstance(normalized, (int, float)):
            values.append((feature["join_key"], feature["name"], float(normalized)))
    return values


def _outlier_payload(metrics_payload: dict[str, Any], metric_id: str, *, limit: int = 5) -> dict[str, Any]:
    values = sorted(_metric_values(metrics_payload, metric_id), key=lambda item: item[2])
    return {
        "metric_id": metric_id,
        "bottom": [
            {"join_key": join_key, "name": name, "normalized_value": value}
            for join_key, name, value in values[:limit]
        ],
        "top": [
            {"join_key": join_key, "name": name, "normalized_value": value}
            for join_key, name, value in values[-limit:][::-1]
        ],
    }


def build_wgi_recipe_payload(generated_at: str) -> dict[str, Any]:
    return {
        "schema_version": 1,
        "recipe_id": "wgi_state_capacity_v1",
        "title": "WGI governance proxy v1 source recipe",
        "source_family": "World Bank Worldwide Governance Indicators",
        "phase": "real_source_cache_v1",
        "source_policy": "real_source_cache_only",
        "generated_at": generated_at,
        "download_policy": {
            "network_allowed": False,
            "default_builder_downloads": False,
            "source_cache_path": WGI_SOURCE_CACHE_RELATIVE_PATH,
        },
        "official_sources": [
            {
                "source_id": "world_bank_wgi_2025_revision",
                "url": WGI_SOURCE_URL,
                "source_package_url": WGI_SOURCE_PACKAGE_URL,
                "release": WGI_SOURCE_TITLE,
                "version": WGI_SOURCE_VERSION,
                "periodicity": "annual",
                "temporal_coverage": "1996-2024",
                "license": WGI_LICENSE,
                "citation": WGI_CITATION,
                "selection_rule": "Use 2024 country/economy rows from official ge and rl sheets; derive a project proxy only when both scores exist.",
            }
        ],
        "metric_selection": {
            "year": WGI_SELECTED_YEAR,
            "source_sheets": {"ge": "government_effectiveness", "rl": "rule_of_law"},
            "official_dimensions": WGI_OFFICIAL_DIMENSIONS,
            "metrics": list(WGI_METRIC_IDS),
            "project_defined_metrics": {
                WGI_COMPOSITE_METRIC_ID: "Mean of Government Effectiveness and Rule of Law when both source scores are observed."
            },
        },
        "join_key_policy": {
            "join_key_type": "iso_a3",
            "special_code_map": SOURCE_CODE_TO_ISO_A3,
            "aggregate_codes_are_dropped": sorted(WGI_AGGREGATE_SOURCE_CODES),
            "unmatched_rows_are_audited": True,
            "name_fuzzy_matching": False,
        },
        "missing_value_policy": "Missing or unavailable source values stay null and carry source_gap or partial_source_gap status.",
        "uncertainty_policy": "Preserve WGI number of sources, standard errors, and 90% confidence intervals for source metrics; do not infer composite uncertainty.",
    }


def build_manifest_payload(
    *,
    metrics_payload: dict[str, Any],
    coverage_counts: dict[str, int],
    signature: WgiSourceSignature,
    generated_at: str,
    accessed_at: str,
) -> dict[str, Any]:
    return {
        "schema_version": 1,
        "layer_id": WGI_LAYER_ID,
        "theme": "political",
        "title": "WGI Governance Proxy",
        "description": "World Bank WGI 2024 admin0 Government Effectiveness and Rule of Law scores with a project-defined state-capacity proxy.",
        "geometry_kind": "admin0",
        "metric_ids": list(WGI_METRIC_IDS),
        "period": {
            "kind": "year",
            "year": WGI_SELECTED_YEAR,
            "label": "2024",
            "basis": "WGI 2025 Revision country/economy rows.",
        },
        "coverage_scope": {
            "geography_level": "admin0",
            "join_key_type": "iso_a3",
            "feature_count": coverage_counts["features"],
            "source_year": WGI_SELECTED_YEAR,
        },
        "source_policy": "real_source_cache_only",
        "status": "experimental",
        "paths": {
            "metrics": data_url(WGI_METRICS_RELATIVE_PATH),
            "build_audit": data_url(WGI_AUDIT_RELATIVE_PATH),
            "source_recipes": [data_url(WGI_RECIPE_RELATIVE_PATH)],
        },
        "provenance": [
            {
                "source_id": "world_bank_wgi_2025_revision",
                "name": "Worldwide Governance Indicators",
                "url": WGI_SOURCE_URL,
                "source_package_url": WGI_SOURCE_PACKAGE_URL,
                "release": WGI_SOURCE_TITLE,
                "version": WGI_SOURCE_VERSION,
                "published_at": WGI_PACKAGE_PUBLISHED_AT,
                "metadata_updated_at": WGI_METADATA_UPDATED_AT,
                "accessed_at": accessed_at,
                "temporal_coverage": "1996-2024",
                "selected_year": WGI_SELECTED_YEAR,
                "license": WGI_LICENSE,
                "citation": WGI_CITATION,
                "selection_rule": "Use 2024 rows from official ge and rl sheets; mean the two 0-100 scores only for the project-defined proxy when both source scores are present.",
                "source_cache_path": signature.repo_relative_path,
                "source_sha256": signature.sha256,
                "source_size_bytes": signature.size_bytes,
                "official_dimensions": WGI_OFFICIAL_DIMENSIONS,
                "uncertainty_fields_preserved": [
                    "number_of_sources",
                    "score_standard_error",
                    "score_confidence_interval_90",
                    "estimate",
                    "estimate_standard_error",
                    "estimate_confidence_interval_90",
                ],
            }
        ],
        "license": {
            "fixture_data": "none",
            "source_metadata": [
                WGI_LICENSE,
                "World Bank Data Catalog metadata last updated 2026-03-18.",
            ],
            "attribution_required": True,
        },
        "normalization": {
            "method": "source_score_0_100_passthrough",
            "range": [0, 100],
            "composite": "project_defined_proxy_mean(government_effectiveness, rule_of_law) when both are observed",
        },
        "feature_counts": coverage_counts,
        "build_audit_path": data_url(WGI_AUDIT_RELATIVE_PATH),
        "generated_at": generated_at,
        "build_command": "python tools/build_thematic_layers.py --include-wgi-real",
        "runtime_consumer": {
            "status": "catalog_only",
            "entry": "thematic_layer_catalog",
            "supports_main_map_render": False,
        },
        "limitations": [
            "WGI scores are country/economy-level governance indicators and are not subnational topology measures.",
            "The state-capacity proxy is a project-defined two-indicator mean, not an official World Bank index or rating.",
            "WGI includes uncertainty; source metric standard errors and 90% confidence intervals are preserved in metrics.admin0.json.",
            "Rows without explicit join keys stay out of the metric payload and are reported in build_audit.",
            "The layer is catalog-only until a later UI/runtime rendering phase accepts it.",
        ],
    }


def build_audit_payload(
    *,
    metrics_payload: dict[str, Any],
    coverage_counts: dict[str, int],
    signature: WgiSourceSignature,
    aggregate_rows: list[DroppedSourceRow],
    unmatched_rows: list[DroppedSourceRow],
    generated_at: str,
    accessed_at: str,
) -> dict[str, Any]:
    warnings: list[str] = []
    if unmatched_rows:
        warnings.append(f"{len(unmatched_rows)} WGI source rows were not mapped to ISO_A3 join keys.")
    if aggregate_rows:
        warnings.append(f"{len(aggregate_rows)} WGI aggregate rows were dropped before metrics generation.")
    if coverage_counts["partial"]:
        warnings.append(f"{coverage_counts['partial']} features have partial source coverage.")
    return {
        "schema_version": 1,
        "layer_id": WGI_LAYER_ID,
        "generated_at": generated_at,
        "builder": {
            "tool": "tools/build_thematic_layers.py",
            "command": "python tools/build_thematic_layers.py --include-wgi-real",
            "ingest_module": "map_builder.thematic_wgi_ingest",
        },
        "source_inputs": [
            {
                "recipe_path": data_url(WGI_RECIPE_RELATIVE_PATH),
                "source_policy": "real_source_cache_only",
                "status": "local_cache_observed",
                "title": WGI_SOURCE_TITLE,
                "version": WGI_SOURCE_VERSION,
                "year": WGI_SELECTED_YEAR,
                "license": WGI_LICENSE,
                "source_cache_path": signature.repo_relative_path,
                "source_sha256": signature.sha256,
                "source_size_bytes": signature.size_bytes,
                "accessed_at": accessed_at,
            }
        ],
        "coverage_summary": {
            **coverage_counts,
            "selected_year": WGI_SELECTED_YEAR,
            "join_key_type": "iso_a3",
            "metric_count": len(WGI_METRIC_IDS),
            "source_rows_unmatched": len(unmatched_rows),
            "source_rows_dropped_aggregate": len(aggregate_rows),
        },
        "missing_join_keys": [],
        "unmatched_source_rows": [_source_row_payload(row) for row in unmatched_rows],
        "dropped_aggregate_rows": [_source_row_payload(row) for row in aggregate_rows],
        "outliers": [_outlier_payload(metrics_payload, metric_id) for metric_id in WGI_METRIC_IDS],
        "normalization_summary": {
            "government_effectiveness": "WGI Governance score (0-100) passthrough.",
            "rule_of_law": "WGI Governance score (0-100) passthrough.",
            "state_capacity_composite": "Project-defined mean of government effectiveness and rule of law when both are observed.",
            "missing_value_policy": "Null raw_value and normalized_value for source_gap or partial_source_gap.",
            "uncertainty_policy": "Preserve source metric standard errors and 90% confidence intervals; do not infer composite uncertainty.",
        },
        "license_summary": {
            "fixture_data": "none",
            "source_metadata": [
                WGI_LICENSE,
                "World Bank WGI Data Catalog entry and package metadata are recorded in manifest provenance.",
            ],
            "attribution_required": True,
        },
        "warnings": warnings,
        "fixture_notice": {
            "enabled": False,
            "reason": "Layer metrics are derived from a local real-source WGI cache.",
        },
    }


def build_wgi_real_source_payloads(
    source_path: Path = DEFAULT_WGI_SOURCE_CACHE_PATH,
    *,
    generated_at: str = DEFAULT_GENERATED_AT,
    accessed_at: str | None = None,
) -> dict[str, dict[str, Any]]:
    accessed_at = accessed_at or generated_at
    signature = source_signature(source_path)
    observations, aggregate_rows, unmatched_rows = load_wgi_observations(signature.path)
    metrics_payload = build_admin_metrics_payload(observations)
    if not metrics_payload["features"]:
        raise ValueError("WGI source cache produced no admin0 features for the selected year")
    coverage_counts = coverage_counts_for_admin(metrics_payload)
    return {
        WGI_RECIPE_RELATIVE_PATH: build_wgi_recipe_payload(generated_at),
        WGI_METRICS_RELATIVE_PATH: metrics_payload,
        WGI_MANIFEST_RELATIVE_PATH: build_manifest_payload(
            metrics_payload=metrics_payload,
            coverage_counts=coverage_counts,
            signature=signature,
            generated_at=generated_at,
            accessed_at=accessed_at,
        ),
        WGI_AUDIT_RELATIVE_PATH: build_audit_payload(
            metrics_payload=metrics_payload,
            coverage_counts=coverage_counts,
            signature=signature,
            aggregate_rows=aggregate_rows,
            unmatched_rows=unmatched_rows,
            generated_at=generated_at,
            accessed_at=accessed_at,
        ),
    }
